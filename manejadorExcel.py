import openpyxl
from openpyxl  import load_workbook
from openpyxl import Workbook

def CrearArchivo(path):
    libro =  Workbook()
    libro.save(path)

def EditarAarchivo(path,datos):
    libro = load_workbook(path)
    hoja = libro.active
    for row in datos:
     hoja.append(row)    
    libro.save(path)

def leerColumna (path,hoja, columna):
   libro = openpyxl.load_workbook(path)
   hoja = libro.get_sheet_by_name(hoja)
   filas = hoja.max_row
   datos = []
   for i in range (1 ,filas + 1):
      celda = hoja.cell(row = i, column = columna)
      datos.append(celda.value)
   return datos

def listaHojas(path):
   libro = openpyxl.load_workbook(path)
   hojas = libro.get_sheet_names()
   return hojas
